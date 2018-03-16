from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def write_indeed():
    import sqlite3
    conn = sqlite3.connect('indeed.db')
    cur = conn.cursor()
    wb = Workbook()
    ws = wb.active
    row = 1
    ws[get_column_letter(1) + str(row)] = 'Title'
    ws[get_column_letter(2) + str(row)] = 'Description'
    ws[get_column_letter(3) + str(row)] = 'Category'
    ws[get_column_letter(4) + str(row)] = 'Link'
    ws[get_column_letter(5) + str(row)] = 'Keyword'
    row += 1
    for line in cur.execute('select * from PRODUCT'):
        counter = 0
        for field in line:
            ws[get_column_letter(counter+1) + str(row)] = field
            counter += 1
        row += 1
    wb.save('indeed_keywords' + '.xlsx')

write_indeed()