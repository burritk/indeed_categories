import traceback
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import NoSuchElementException

from pyscraper.selenium_utils import get_headed_driver, wait_for_xpath, get_headless_driver

driver = get_headed_driver()
driver.get('https://www.indeed.com/q-Full-Time-l-California-jobs.html')

# listings = driver.find_elements_by_class_name('row result clickcard')

hrefs = []
while len(hrefs) < 3:
    listings = driver.find_elements_by_xpath('//div[@class="row result clickcard"]')
    for listing in listings:
        # company = listing.find_element_by_class_name('company').find_element_by_tag_name('a').get_attribute('href')
        try:
            company = listing.find_element_by_xpath('./span[1]/a').get_attribute('href')
            if company in hrefs:
                continue
            hrefs.append(company)
            print company
        except NoSuchElementException:
            # print 'No Company Link'
            continue
    try:
        last_item = driver.find_element_by_xpath('//div[@class="lastRow row result clickcard"]')
        company = last_item.find_element_by_xpath('./span[1]/a').get_attribute('href')
        if company not in hrefs:
            hrefs.append(company)
            print company
    except NoSuchElementException:
        pass
    page_links = driver.find_element_by_class_name('pagination').find_elements_by_tag_name('a')
    # page_links = driver.find_elements_by_xpath('//*[@id="resultsCol"]/div[13]/a')
    next_button = page_links[-1]
    driver.get(next_button.get_attribute('href'))

wb = Workbook()
ws = wb.active
row = 1
headers = ['Name',
           'Rating',
           'Link',
           'About',
           'Employees',
           'Headquarters',
           'Industry',
           'Total Jobs',
           'Jobs by Location',
           'Jobs',
           'Company Links'
           ]

# final_values.append(headers)
for index, header in enumerate(headers):
    ws[get_column_letter(index+1) + str(row)] = header
row += 1

for index, href in enumerate(hrefs):
    driver.get(href)
    company_name = driver.find_element_by_xpath('//*[@id="cmp-name-and-rating"]/div[1]').text
    company_rating = driver.find_element_by_xpath('//*[@id="cmp-header-rating"]/span').text

    company_hq = ''
    company_employees = ''
    company_industry = ''
    company_links = ''
    try:
        company_info = driver.find_element_by_xpath('//*[@id="cmp-company-details-sidebar"]')
        info_lines = company_info.find_elements_by_xpath('./*')
        for index, line in enumerate(info_lines):
            if 'Headquarters' in line.text:
                company_hq = info_lines[index+1].text
                continue
            if 'Employees' in line.text:
                company_employees = info_lines[index+1].text
                continue
            if 'Industry' in line.text:
                company_industry = info_lines[index+1].text
                continue
            if 'Links' in line.text:
                for link in info_lines[index+1].find_elements_by_tag_name('a'):
                    company_links += link.text + '(' + link.get_attribute('href') + ')\n'
    except:
        pass

    company_jobs = []
    jobs = driver.find_element_by_xpath('//*[@id="cmp-jobs-container"]')
    for ul in jobs.find_elements_by_xpath('./ul'):
        for li in ul.find_elements_by_xpath('./li'):
            company_jobs.append(li.text)

    company_jobs_location = '\n'.join(company_jobs)
    total_jobs = driver.find_element_by_xpath('//*[@id="cmp-jobs"]/div[2]/a').text
    company_total_jobs = filter(lambda x: x.isdigit(), total_jobs)
    try:
        company_about = driver.find_element_by_xpath('//*[@id="cmp-short-description"]').text
    except:
        company_about = ''
    print 'h'

    jobs_tab = driver.find_element_by_xpath('//*[@id="cmp-menu-container"]/ul/li[5]/a')
    driver.get(jobs_tab.get_attribute('href'))
    all_jobs = driver.find_element_by_xpath('//*[@id="cmp-jobs"]/div[2]/div[2]/a')
    driver.get(all_jobs.get_attribute('href'))

    company_jobs = []

    while len(company_jobs) < 200:
        page = driver.find_element_by_xpath('//*[@id="cmp-jobs"]/ul')
        for item in page.find_elements_by_xpath('./li'):
            '//*[@id="cmp-jobs"]/ul/li[1]/div/h3'
            job_title = item.find_element_by_xpath('./div/h3').text
            job_location = item.find_element_by_xpath('./div/div[1]').text
            company_jobs.append('{0} ({1})'.format(job_title.encode('utf-8'), job_location.encode('utf-8')))
        try:
            next_button = driver.find_elements_by_xpath('//*[@id="cmp-jobs"]/div[3]/a')[-1]
            if 'next' not in next_button.text.lower():
                break
        except:
            break
        driver.get(next_button.get_attribute('href'))

    company_jobs = '\n'.join(company_jobs)
    # company_job_number = driver.find_element_by_xpath('//*[@id="cmp-jobs"]/div[2]/span').text
    data = [company_name,
               company_rating,
               href.encode('utf-8'),
               company_about,
               company_employees,
               company_hq,
               company_industry,
               company_total_jobs,
               company_jobs_location,
               company_jobs,
               company_links]

    for index, datum in enumerate(data):
        ws[get_column_letter(index+1) + str(row)] = datum
    row += 1
wb.save('indeed.xlsx')
driver.close()
